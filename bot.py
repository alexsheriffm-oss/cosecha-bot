import os
import json
import logging
from datetime import datetime
from zoneinfo import ZoneInfo
import gspread
from google.oauth2.service_account import Credentials
from telegram import Update
from telegram.ext import (
    Application, CommandHandler, ContextTypes, JobQueue
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN  = os.environ["BOT_TOKEN"]
SHEET_ID   = os.environ["SHEET_ID"]
GROUP_ID   = int(os.environ.get("GROUP_ID", "-1001001003779316015"))
TZ         = ZoneInfo("America/La_Paz")

def get_sheet():
    creds_dict = json.loads(os.environ["GOOGLE_CREDS_JSON"])
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds  = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    client = gspread.authorize(creds)
    sh     = client.open_by_key(SHEET_ID)
    try:
        ws = sh.worksheet("Registros")
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet("Registros", rows=5000, cols=7)
        ws.append_row(["user_id", "nombre", "fecha", "tipo", "modalidad", "hora", "horas_dia"])
    return ws

def now_lp():
    return datetime.now(TZ)

def fmt_date(dt): return dt.strftime("%d/%m/%Y")
def fmt_time(dt): return dt.strftime("%H:%M")

def fmt_dur(secs):
    h = int(secs // 3600)
    m = int((secs % 3600) // 60)
    return f"{h}h {m:02d}min"

def get_user_rows(ws, user_id, fecha=None):
    all_rows = ws.get_all_records()
    return [
        r for r in all_rows
        if str(r["user_id"]) == str(user_id)
        and (fecha is None or r["fecha"] == fecha)
    ]

def calc_hours(rows):
    entradas = [r for r in rows if r["tipo"] == "ENTRO"]
    salidas  = [r for r in rows if r["tipo"] == "SALGO"]
    total    = 0.0
    for i, e in enumerate(entradas):
        try:
            fecha_str = e["fecha"]
            t_e = datetime.strptime(f"{fecha_str} {e['hora']}", "%d/%m/%Y %H:%M").replace(tzinfo=TZ)
            if i < len(salidas):
                t_s = datetime.strptime(f"{fecha_str} {salidas[i]['hora']}", "%d/%m/%Y %H:%M").replace(tzinfo=TZ)
                diff = (t_s - t_e).total_seconds()
            else:
                t_now = now_lp().replace(second=0, microsecond=0)
                diff  = (t_now - t_e).total_seconds()
            if 0 < diff <= 57600:
                total += diff
        except Exception:
            continue
    return total

def is_working(rows):
    entradas = [r for r in rows if r["tipo"] == "ENTRO"]
    salidas  = [r for r in rows if r["tipo"] == "SALGO"]
    return len(entradas) > len(salidas)

def get_modalidad_hoy(rows):
    entradas = [r for r in rows if r["tipo"] == "ENTRO"]
    if entradas:
        modalidad = str(entradas[-1].get("modalidad", "")).strip().lower()
        return modalidad if modalidad in ("virtual", "presencial") else "presencial"
    return None

async def entro(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user  = update.effective_user
    now   = now_lp()
    fecha = fmt_date(now)
    hora  = fmt_time(now)
    ws    = get_sheet()
    rows  = get_user_rows(ws, user.id, fecha)

    if is_working(rows):
        modalidad = get_modalidad_hoy(rows)
        if modalidad == "virtual":
            await update.message.reply_text(
                f"⚠️ {user.first_name}, ya entraste como *virtual* hoy.\n"
                f"No puedes registrar entrada presencial el mismo día.\n"
                f"Usa /salgo cuando termines.",
                parse_mode="Markdown"
            )
        else:
            hora_entrada = [r for r in rows if r["tipo"] == "ENTRO"][-1]["hora"]
            await update.message.reply_text(
                f"⚠️ {user.first_name}, ya estás trabajando desde las *{hora_entrada}*.\n"
                f"Usa /salir cuando termines.",
                parse_mode="Markdown"
            )
        return

    ws.append_row([str(user.id), user.full_name, fecha, "ENTRO", "presencial", hora, ""])
    await update.message.reply_text(
        f"✅ *{user.full_name}* entró a las *{hora}*\n"
        f"🏢 Modalidad: presencial\n"
        f"📅 {fecha}",
        parse_mode="Markdown"
    )

async def entrovirtual(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user  = update.effective_user
    now   = now_lp()
    fecha = fmt_date(now)
    hora  = fmt_time(now)
    ws    = get_sheet()
    rows  = get_user_rows(ws, user.id, fecha)

    if is_working(rows):
        modalidad = get_modalidad_hoy(rows)
        if modalidad == "presencial":
            await update.message.reply_text(
                f"⚠️ {user.first_name}, ya entraste de forma *presencial* hoy.\n"
                f"No puedes registrar entrada virtual el mismo día.\n"
                f"Usa /salir cuando termines.",
                parse_mode="Markdown"
            )
        else:
            hora_entrada = [r for r in rows if r["tipo"] == "ENTRO"][-1]["hora"]
            await update.message.reply_text(
                f"⚠️ {user.first_name}, ya estás trabajando (virtual) desde las *{hora_entrada}*.\n"
                f"Usa /salir cuando termines.",
                parse_mode="Markdown"
            )
        return

    ws.append_row([str(user.id), user.full_name, fecha, "ENTRO", "virtual", hora, ""])
    await update.message.reply_text(
        f"✅ *{user.full_name}* entró a las *{hora}*\n"
        f"💻 Modalidad: virtual\n"
        f"📅 {fecha}",
        parse_mode="Markdown"
    )

async def salgo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    user  = update.effective_user
    now   = now_lp()
    fecha = fmt_date(now)
    hora  = fmt_time(now)
    ws    = get_sheet()
    rows  = get_user_rows(ws, user.id, fecha)

    if not is_working(rows):
        await update.message.reply_text(
            f"⚠️ {user.first_name}, no tienes una entrada activa hoy."
        )
        return

    modalidad = get_modalidad_hoy(rows)
    ws.append_row([str(user.id), user.full_name, fecha, "SALGO", modalidad, hora, ""])

    rows_act  = get_user_rows(ws, user.id, fecha)
    horas     = calc_hours(rows_act)
    horas_txt = fmt_dur(horas)

    all_data = ws.get_all_values()
    for i in range(len(all_data) - 1, 0, -1):
        if all_data[i][0] == str(user.id) and all_data[i][3] == "SALGO" and all_data[i][6] == "":
            ws.update_cell(i + 1, 7, horas_txt)
            break

    icono = "🏢" if modalidad == "presencial" else "💻"
    await update.message.reply_text(
        f"🔴 *{user.full_name}* salió a las *{hora}*\n"
        f"{icono} Modalidad: {modalidad}\n"
        f"⏱ Hoy trabajó: *{horas_txt}*",
        parse_mode="Markdown"
    )

async def equipo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    now      = now_lp()
    fecha    = fmt_date(now)
    ws       = get_sheet()
    all_rows = ws.get_all_records()

    from collections import defaultdict
    usuarios = defaultdict(list)
    for r in all_rows:
        uid = str(r["user_id"])
        if uid:
            usuarios[uid].append(r)

    if not usuarios:
        await update.message.reply_text("📭 Aún no hay registros.")
        return

    presencial_on = []
    virtual_on    = []
    fuera         = []

    for uid, rows in usuarios.items():
        nombre    = rows[-1]["nombre"]
        rows_hoy  = [r for r in rows if r["fecha"] == fecha]
        working   = is_working(rows_hoy)
        horas     = calc_hours(rows_hoy)
        modalidad = get_modalidad_hoy(rows_hoy)

        if working:
            entrada_hora = [r for r in rows_hoy if r["tipo"] == "ENTRO"][-1]["hora"]
            dato = (nombre, fmt_dur(horas), entrada_hora)
            if modalidad and modalidad.strip().lower() == "virtual":
                virtual_on.append(dato)
            else:
                presencial_on.append(dato)
        else:
            if rows_hoy:
                ultima = rows_hoy[-1]["hora"]
                fuera.append((nombre, fmt_dur(horas), f"salió {ultima}"))
            else:
                fuera.append((nombre, "0h 00min", "no registró hoy"))

    lineas = [
        f"🌱 *Biométrico — Cosecha Colectiva*",
        f"📅 {fecha}  •  🕐 {fmt_time(now)}",
        "━━━━━━━━━━━━━━━━",
    ]

    if presencial_on:
        lineas.append(f"🏢 *PRESENCIAL ({len(presencial_on)})*")
        for nombre, horas_txt, desde in sorted(presencial_on):
            lineas.append(f"  🟢 {nombre}")
            lineas.append(f"       ⏱ {horas_txt} · desde {desde}")
    else:
        lineas.append("🏢 *PRESENCIAL* — nadie en oficina")

    lineas.append("━━━━━━━━━━━━━━━━")

    if virtual_on:
        lineas.append(f"💻 *VIRTUAL ({len(virtual_on)})*")
        for nombre, horas_txt, desde in sorted(virtual_on):
            lineas.append(f"  🟢 {nombre}")
            lineas.append(f"       ⏱ {horas_txt} · desde {desde}")
    else:
        lineas.append("💻 *VIRTUAL* — nadie conectado")

    lineas.append("━━━━━━━━━━━━━━━━")

    if fuera:
        lineas.append(f"⚪ *FUERA DE TURNO ({len(fuera)})*")
        for nombre, horas_txt, estado in sorted(fuera):
            lineas.append(f"  👤 {nombre}  —  {estado}")

    total_on = len(presencial_on) + len(virtual_on)
    lineas.append("━━━━━━━━━━━━━━━━")
    lineas.append(f"👥 *{total_on} trabajando* de {len(usuarios)} en el equipo")

    await update.message.reply_text("\n".join(lineas), parse_mode="Markdown")

async def reporte(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    now = now_lp()
    if ctx.args:
        try:
            mes, anio = ctx.args[0].split("/")
            mes, anio = int(mes), int(anio)
        except Exception:
            await update.message.reply_text("❌ Formato: /reporte MM/AAAA  (ej: /reporte 05/2026)")
            return
    else:
        mes, anio = now.month, now.year

    await update.message.reply_text(f"⏳ Generando reporte {mes:02d}/{anio}...")

    ws        = get_sheet()
    all_rows  = ws.get_all_records()
    mes_str   = f"{mes:02d}"
    anio_str  = str(anio)

    filtrados = [
        r for r in all_rows
        if r.get("fecha") and str(r["fecha"])[3:5] == mes_str and str(r["fecha"])[6:] == anio_str
    ]

    if not filtrados:
        await update.message.reply_text(f"📭 No hay registros para {mes:02d}/{anio}.")
        return

    from collections import defaultdict
    resumen  = defaultdict(lambda: defaultdict(lambda: {"secs": 0.0, "modalidad": "presencial"}))
    usuarios = defaultdict(list)
    nombres  = {}

    for r in filtrados:
        uid = str(r["user_id"])
        usuarios[uid].append(r)
        nombres[uid] = r["nombre"]

    for uid, rows in usuarios.items():
        fechas = set(r["fecha"] for r in rows)
        for fecha in fechas:
            rows_dia  = [r for r in rows if r["fecha"] == fecha]
            secs      = calc_hours(rows_dia)
            modalidad = get_modalidad_hoy(rows_dia) or "presencial"
            resumen[nombres[uid]][fecha] = {"secs": secs, "modalidad": modalidad}

    wb    = openpyxl.Workbook()
    ws_xl = wb.active
    ws_xl.title = f"Reporte {mes:02d}-{anio}"

    V_OSC  = "1B4332"
    V_MED  = "40916C"
    AZUL   = "185FA5"
    BLANCO = "FFFFFF"
    GRIS   = "F8F9FA"

    ws_xl.merge_cells("A1:F1")
    t = ws_xl["A1"]
    t.value     = f"COSECHA COLECTIVA — Horas trabajadas {mes:02d}/{anio}"
    t.font      = Font(bold=True, size=14, color=BLANCO)
    t.fill      = PatternFill("solid", fgColor=V_OSC)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_xl.row_dimensions[1].height = 30

    for col, h in enumerate(["Nombre", "Fecha", "Modalidad", "Horas", "Minutos", ""], 1):
        c = ws_xl.cell(row=2, column=col, value=h)
        c.font      = Font(bold=True, color=BLANCO, size=11)
        c.fill      = PatternFill("solid", fgColor=V_MED)
        c.alignment = Alignment(horizontal="center")

    fila = 3
    totales = {}
    for nombre in sorted(resumen.keys()):
        total_secs = 0.0
        for fecha in sorted(resumen[nombre].keys()):
            dato      = resumen[nombre][fecha]
            secs      = dato["secs"]
            modalidad = dato["modalidad"]
            total_secs += secs
            h = int(secs // 3600)
            m = int((secs % 3600) // 60)

            bg = BLANCO if fila % 2 == 0 else GRIS
            datos = [nombre, fecha, modalidad.upper(), f"{h}h {m:02d}min", int(secs // 60), ""]
            for col, val in enumerate(datos, 1):
                cell = ws_xl.cell(row=fila, column=col, value=val)
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.alignment = Alignment(horizontal="left" if col == 1 else "center")
                if col == 3:
                    cell.font = Font(color=AZUL if modalidad == "virtual" else V_OSC, bold=True, size=10)
            fila += 1

        totales[nombre] = total_secs
        h = int(total_secs // 3600)
        m = int((total_secs % 3600) // 60)
        ws_xl.merge_cells(f"A{fila}:B{fila}")
        for col in range(1, 7):
            c = ws_xl.cell(row=fila, column=col)
            c.fill = PatternFill("solid", fgColor=V_MED)
            c.font = Font(bold=True, color=BLANCO)
            c.alignment = Alignment(horizontal="center")
        ws_xl.cell(row=fila, column=1, value=f"TOTAL  {nombre.upper()}")
        ws_xl.cell(row=fila, column=4, value=f"{h}h {m:02d}min")
        ws_xl.cell(row=fila, column=5, value=int(total_secs // 60))
        fila += 2

    gran_total = sum(totales.values())
    h = int(gran_total // 3600)
    m = int((gran_total % 3600) // 60)
    ws_xl.merge_cells(f"A{fila}:B{fila}")
    for col in range(1, 7):
        c = ws_xl.cell(row=fila, column=col)
        c.fill = PatternFill("solid", fgColor=V_OSC)
        c.font = Font(bold=True, size=12, color=BLANCO)
        c.alignment = Alignment(horizontal="center")
    ws_xl.cell(row=fila, column=1, value="TOTAL EQUIPO")
    ws_xl.cell(row=fila, column=4, value=f"{h}h {m:02d}min")
    ws_xl.cell(row=fila, column=5, value=int(gran_total // 60))

    ws_xl.column_dimensions["A"].width = 28
    ws_xl.column_dimensions["B"].width = 14
    ws_xl.column_dimensions["C"].width = 14
    ws_xl.column_dimensions["D"].width = 14
    ws_xl.column_dimensions["E"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await update.message.reply_document(
        document=buf,
        filename=f"CosechaColectiva_Horas_{mes:02d}_{anio}.xlsx",
        caption=f"📊 Reporte {mes:02d}/{anio} — Cosecha Colectiva\n🏢 Presencial + 💻 Virtual"
    )

async def auto_desconectar(ctx: ContextTypes.DEFAULT_TYPE):
    now   = now_lp()
    fecha = fmt_date(now)
    ws    = get_sheet()
    all_rows = ws.get_all_records()

    from collections import defaultdict
    usuarios = defaultdict(list)
    for r in all_rows:
        uid = str(r["user_id"])
        if uid:
            usuarios[uid].append(r)

    cerrados = []
    for uid, rows in usuarios.items():
        rows_hoy = [r for r in rows if r["fecha"] == fecha]
        if is_working(rows_hoy):
            nombre    = rows_hoy[-1]["nombre"]
            modalidad = get_modalidad_hoy(rows_hoy)
            horas     = calc_hours(rows_hoy)
            ws.append_row([uid, nombre, fecha, "SALGO", modalidad, "23:59", fmt_dur(horas)])
            cerrados.append(nombre)

    if cerrados and ctx.job.data:
        nombres_txt = "\n".join(f"  • {n}" for n in cerrados)
        await ctx.bot.send_message(
            chat_id=ctx.job.data,
            text=(
                f"🌙 *Cierre automático de medianoche*\n"
                f"Se cerró la sesión de:\n{nombres_txt}\n\n"
                f"_Recuerden registrar su entrada mañana._ 🌱"
            ),
            parse_mode="Markdown"
        )

def main():
    app = Application.builder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("entro",        entro))
    app.add_handler(CommandHandler("entrovirtual", entrovirtual))
    app.add_handler(CommandHandler("salgo",        salir))
    app.add_handler(CommandHandler("equipo",       equipo))
    app.add_handler(CommandHandler("reporte",      reporte))

    job_queue = app.job_queue
    job_queue.run_daily(
        auto_desconectar,
        time=datetime.strptime("23:59", "%H:%M").replace(tzinfo=TZ).timetz(),
        data=GROUP_ID,
        name="medianoche"
    )

    logger.info("🌱 Wason v4 arrancando — modo grupo...")
    app.run_polling(drop_pending_updates=True)

if __name__ == "__main__":
    main()
